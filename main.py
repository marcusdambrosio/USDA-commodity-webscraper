import pandas as pd
import numpy as np
import os, sys, time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import dropdown_elements
import datetime as dt
from openpyxl import load_workbook
import re
import pyexcel
import math
from crop_formatting import *
import tkinter
from tkinter import ttk

'''constants'''
global daysBack
global fileDate
nameConverter = {'AJ':'ATLANTA',
                 'BP':'BALTIMORE',
                 'BH':'BOSTON',
                 'HX':'CHICAGO',
                 'CA':'COLUMBIA',
                 'DA':'DALLAS',
                 'DU':'DETROIT',
                 'GU':'GUADALAJARA',
                 'HC':'LOS ANGELES',
                 'MX':'MEXICO CITY',
                 'MH':'MIAMI',
                 'MT':'MONTERREY',
                 'NX':'NEW YORK',
                 'NA':'PHILADELPHIA',
                 'RO':'ROTTERDAM',
                 'SX':'SAN FRANCISCO',
                 'TO':'TORONTO, ONT'}

downloadNums = {}
for i, item in enumerate(nameConverter.values()):
    downloadNums[i] = item

def download():
    '''
    Clears duplicate files and download all cropdata
    :return: locally downloaded files
    '''
    URL = 'https://www.marketnews.usda.gov/mnp/fv-report-config-step2?repType=wiz&type=termPrice&locChoose=location&commodityClass=allcommodity&run=Run'
    driver = webdriver.Chrome()
    driver.get(URL)
    downloadFiles = [f for f in os.listdir(r'C:\Users\marcu\Downloads') if 'report' in f]
    for badFile in downloadFiles:
        os.remove(os.path.join(r'C:\Users\marcu\Downloads', badFile))
    yearsBack = int(np.floor(daysBack/365)) if daysBack>=365 else 1
    for yearBack in range(yearsBack):
        for state in dropdown_elements.states:
            locDropdown = Select(driver.find_element_by_id('locAbr'))
            locDropdown.select_by_value(state)
            commodityDropdown = Select(driver.find_element_by_id('commAbrfrom'))
            for crop in dropdown_elements.focusedCrops:
                commodityDropdown.select_by_visible_text(crop)
                driver.find_element_by_id('Add').click()

            # #open up date boxes
            driver.find_element_by_id('step3').click()
            time.sleep(1)
            startDate = driver.find_element_by_id('repDate')
            extraDays= 0
            if daysBack<364:
                startD = (dt.datetime.now() - dt.timedelta(days = daysBack)).strftime('%m/%d/%Y')
                endD = dt.datetime.now().strftime('%m/%d/%Y')
            else:
                startD = (dt.datetime.now() - dt.timedelta(days = 364*(yearBack+1) + extraDays)).strftime('%m/%d/%Y')
                endD = (dt.datetime.now() - dt.timedelta(days = 364*yearBack + extraDays)).strftime('%m/%d/%Y')

            startDate.send_keys(startD)
            endDate = driver.find_element_by_id('endDate')
            endDate.send_keys(endD)

            #run query button
            driver.find_element_by_id('Run').click()
            time.sleep(20)
            #download excel button
            section = driver.find_element_by_class_name('LightYellowBackgroundColor')
            links = section.find_elements_by_class_name('BodyTextBlack')[-1]
            atags = links.find_elements_by_tag_name('a')
            href = [a.get_attribute('href') for a in atags if 'excel' in a.get_attribute('href')][0]
            driver.get(href)
            time.sleep(2)
            driver.get(URL)
            time.sleep(2)
            print(f'{nameConverter[state]} finished\n')
    return (dt.datetime.now() - dt.timedelta(days = daysBack)).strftime('%m/%d/%Y')

def multiyear_comp():
    '''
    Combines data for all locations
    :return: consolidated by city csv saved locally
    '''
    downloads = os.listdir(r'C:\Users\marcu\Downloads')
    downloads = [c for c in downloads if 'report' in c and 'xlsx' in c]

    masters = {}
    for report in downloads:
        data = pd.read_excel(os.path.join(r'C:\Users\marcu\Downloads',report))
        cityName=data.loc[0, 'City Name']
        if cityName not in downloadNums.values():
            print(cityName)
            # downloads[list(downloadNums.keys())[-1]+1] = cityName
        if cityName in masters.keys():
            k = [key for key, value in downloadNums.items() if value == cityName][0]
            masters[k] = masters[k].append(data)
        else:
            k = [key for key, value in downloadNums.items() if value == cityName][0]
            masters[k] = data

    for d in downloads:
        os.remove(os.path.join(r'C:\Users\marcu\Downloads',d))

    for k , data in masters.items():
        if k == list(masters.keys())[0]:
            data.to_excel(r'C:\Users\marcu\Downloads\report.xlsx', index = False)
        else:
            data.to_excel(r'C:\Users\marcu\Downloads\report '+ f'({k}).xlsx', index = False)

def compilation():
    '''
    Combines crop data to multiyear xcel workbook
    :return:
    '''
    if fileDate.replace("/", "-") not in os.listdir('downloads'):
        os.mkdir('downloads/' + fileDate.replace("/", "-"))
    downloads = os.listdir(r'C:\Users\marcu\Downloads')
    downloads = [c for c in downloads if 'report' in c]
    try:
        for d in downloads:
            dat = pd.read_html(os.path.join(r'C:\Users\marcu\Downloads', d))[0]
            dat.to_excel(os.path.join(r'C:\Users\marcu\Downloads', d.split('.')[0]+'.xlsx'))
    except:
        print('Xlsx already created.')

    multiyear_comp()
    print('comp done')
    downloads = os.listdir(r'C:\Users\marcu\Downloads')
    downloads = [c for c in downloads if 'report' in c and 'xlsx' in c]

    first = [f for f in downloads if re.sub("[^0-9]", "", f) == ''][0]
    downloads.remove(first)

    destBook = load_workbook(os.path.join(r'C:\Users\marcu\Downloads', first))
    for pageNum, report in enumerate(downloads):
        print(report)
        convKey = int(re.sub("[^0-9]", "", report))
        destBook.create_sheet(downloadNums[convKey])
        destBook.worksheets[0].title = downloadNums[0]
        destSheet = destBook[downloadNums[convKey]]

        sourceBook = load_workbook(os.path.join(r'C:\Users\marcu\Downloads',report))
        sourceSheet = sourceBook.worksheets[0]

        mr = sourceSheet.max_row
        mc = sourceSheet.max_column

        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                # reading cell value from source excel file
                c = sourceSheet.cell(row=i, column=j)

                # writing the read value to destination excel file
                destSheet.cell(row=i, column=j).value = c.value

    destBook.save(f'downloads/{fileDate.replace("/", "-")}/allcrops.xlsx')

def combination():
    '''
    combines tabs of xcel file to sincle excel sheet
    :return:
    '''
    data = pd.ExcelFile(f'downloads/{fileDate.replace("/", "-")}/allcrops.xlsx')
    master = pd.DataFrame()
    sheets = data.sheet_names
    for sheet in sheets:
        currDf = data.parse(sheet_name = sheet)
        master = master.append(currDf)
    for col in master.columns:
        if 'Unnamed' in col:
            master.drop(col, axis = 1, inplace = True)
    master.to_excel(f'downloads/{fileDate.replace("/", "-")}/combined.xlsx', index = False)

def formatting():
    '''
    manages formatting functions for each individual crop
    :return:
    '''
    data = pd.read_excel(f'downloads/{fileDate.replace("/", "-")}/combined.xlsx')
    crops = ['GRAPES', 'STRAWBERRIES', 'TOMATOES', 'PISTACHIOS', 'ALMONDS', 'ORANGES', 'CHERRIES', 'BLUEBERRIES', 'LETTUCE', 'APRICOTS']
    crops = ['ORANGES', 'CHERRIES', 'BLUEBERRIES','LETTUCE','APRICOTS']
    for row in data.iterrows():
        i, row = row
        if 'TOMATOES' in row['Commodity Name'] and ',' in row['Commodity Name']:
            if type(row['Variety']) != str:
                data.loc[i, 'Variety'] = row['Commodity Name'].split(',')[-1].strip()
            data.loc[i, 'Commodity Name'] = 'TOMATOES'
        elif 'LETTUCE' in row['Commodity Name'] and ',' in row['Commodity Name']:
            if type(row['Variety']) != str:
                data.loc[i, 'Variety'] = row['Commodity Name'].split(',')[-1].strip()
            data.loc[i, 'Commodity Name'] = 'LETTUCE'

    dataDict = {}
    for crop in crops:
        dataDict[crop] = data[data['Commodity Name'] == crop]


    give_daysBack(daysBack)
    TOMATOES(dataDict['TOMATOES'])
    ALMONDS(dataDict['ALMONDS'])
    GRAPES(dataDict['GRAPES'])
    PISTACHIOS(dataDict['PISTACHIOS'])
    STRAWBERRIES(dataDict['STRAWBERRIES'])
    # ORANGES(dataDict['ORANGES'])
    # CHERRIES(dataDict['CHERRIES'])
    # APRICOTS(dataDict['APRICOTS'])
    # LETTUCE(dataDict['LETTUCE'])
    # BLUEBERRIES(dataDict['BLUEBERRIES'])


def recombine():
    '''
    recombines the crop list after formatting
    :return:
    '''
    croplist = ['almonds', 'grapes', 'tomatoes', 'strawberries', 'pistachios', 'oranges', 'apricots', 'blueberries',
                'cherries', 'lettuce']
    croplist = ['oranges', 'apricots', 'blueberries', 'cherries', 'lettuce']

    master = pd.DataFrame()
    for crop in croplist:
        currData = pd.read_csv(f'downloads/{fileDate.replace("/", "-")}/{crop}.csv')
        master = master.append(currData)

    print(master.head())

    master.to_csv(f'downloads/{fileDate.replace("/", "-")}/master_for_upload.csv', index=False)

'''main class to manage'''
class Commodities:
    def __init__(self):
        self.cropList = ['ALMONDS', 'APRICOTS', 'BLUEBERRIES', 'CHERRIES', 'GRAPES',
                         'LETTUCE', 'ORANGES', 'PISTACHIOS', 'STRAWBERRIES',' TOMATOES']
        self.root = tkinter.Tk()
        self.root.geometry('800x600')
        self.root.resizable(False ,False)
        self.entryFrame = tkinter.Frame(self.root,highlightbackground='darkgrey', highlightthickness=2)
        self.entryFrame.grid(row = 0, column = 0, sticky = 'ew', pady = 5, padx=5)
        self.cropButtonFrame = tkinter.Frame(self.entryFrame)
        self.cropButtonFrame.grid(row =0, column = 0, sticky = 'ew', pady = 10, padx=5)
        self.optionButtonFrame = tkinter.Frame(self.entryFrame)
        self.optionButtonFrame.grid(row = 0, column = 1, sticky= 'ew', pady=10, padx= 5)

        self.cropButtonLabel = tkinter.Label(self.cropButtonFrame, text='SELECT CROPS', font = ('Helvetico bold', 16), anchor = 'w')
        self.cropButtonLabel.grid(row =0, column=1, sticky = 'w')
        self.bools = {}
        self.buttons = {}
        for i, crop in enumerate(self.cropList):
            i += 1
            self.bools[crop] = tkinter.BooleanVar()
            self.bools[crop].set(False)
            self.buttons[crop] = tkinter.Checkbutton(self.cropButtonFrame, text = crop,
                                                     variable = self.bools[crop], padx=5, pady=5)
            self.buttons[crop].grid(row = i, column = 0, sticky = 'w', padx = 5)


        self.root.mainloop()


'''RUN'''
if __name__ == '__main__':
    daysBack = int(input('Number of days back:')) - 1
    fileDate = (dt.datetime.now() - dt.timedelta(days = daysBack)).strftime('%m/%d/%Y')
    print('Downloads for {daysBack+1} days beginning...\n')
    tic = dt.datetime.now()
    date = download()
    print(f'Downloads finished in {dt.datetime.now() - tic}\n')
    print('Compilation beginning...\n')
    tic = dt.datetime.now()
    compilation()
    print(f'Compilation finished in {dt.datetime.now() - tic}\n')
    print('Combination beginning...\n')
    tic = dt.datetime.now()
    combination()
    print(f'Combination finished in {dt.datetime.now() - tic}\n')
    fileDate='10-11-2011'
    daysBack = 3639
    print('Formatting beginning...\n')
    tic = dt.datetime.now()
    formatting()
    print(f'Formatting finished in {dt.datetime.now() - tic}\n')
    recombine()

    df = pd.read_excel('downloads/10-11-2011/combined.xlsx')
    df = df[df['Commodity Name'] == 'BLUEBERRIES']
    print(df['Unit of Sale'].unique())
    print(df['Package'].unique())
